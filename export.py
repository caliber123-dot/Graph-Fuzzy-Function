from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import json
from PIL import Image as PILImage  # For resizing
import tempfile  # This was missing in your original code
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

def export_images_to_excel(img1, img2, output_excel="output.xlsx", max_width=500, max_height=300,fn_dict='',filetitle=''):
    """
    Export two resized PNG images to an Excel file with proper temp file handling.
    """
    temp_files = []  # To keep track of temporary files
    
    # Define alpha-cuts and function values
    # alpha_cuts = [0.3, 0.4, 0.5]
    # fn_dict2 = {
    #     'fn1': [810.3771, 819.2512, 822.1110],
    #     'fn2': [859.8149, 856.4664, 853.1098],
    #     'fn3': [812.1549, 815.6181, 819.0718],
    #     'fn4': [855.3681, 852.6683, 849.9560]
    # }
    fn_dict_str = fn_dict
    fn_dict = json.loads(fn_dict_str.replace("'", '"'))  # Replace single quotes with double quotes 
    # print(type(fn_dict))
    # print(fn_dict)
    try:
        # Validate input images
        for img_path in [img1, img2]:
            if not os.path.exists(img_path):
                raise FileNotFoundError(f"Image file not found: {img_path}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "BarChart"

        # Function to resize image and return temp path
        def resize_image(img_path):
            img = PILImage.open(img_path)
            img.thumbnail((max_width, max_height))            
            # Create temp file in system's temp directory
            fd, temp_path = tempfile.mkstemp(suffix='.png')
            os.close(fd)  # We only need the path
            img.save(temp_path)
            temp_files.append(temp_path)  # Track for cleanup
            return temp_path

        # Process images
        resized_img1 = resize_image(img1)
        resized_img2 = resize_image(img2)

        # Add images to worksheet
        # ws.add_image(Image(resized_img1), "A24")
        ws.add_image(Image(resized_img2), "A10")

        # Prepare DataFrame for alpha-cuts
        # df = pd.DataFrame({
        #     # 'α': alpha_cuts,
        #     # **({'α': fn_dict['α']} if fn_dict['α'] is None else {"α-α": fn_dict['α-α']}),
        #     'α': fn_dict['α'],
        #     'fn₁ (Y̲,ρ̲)': fn_dict['fn1'],
        #     'fn₂ (Y̅,ρ̲)': fn_dict['fn2'],
        #     'fn₃ (Y̲,ρ̅ )': fn_dict['fn3'],
        #     'fn₄ (Y̅,ρ̅ )': fn_dict['fn4']
        # })
        data = {}
        # Handle α/α-α columns
        # if 'α-α' in fn_dict and fn_dict['α-α']:
        #     data['α-α'] = fn_dict['α-α']
        # elif 'α' in fn_dict and fn_dict['α'] is not None:
        #     data['α'] = fn_dict['α']
        # Check if 'α' exists and is not None, else use 'α-α'
        cutval = ""
        if 'α' in fn_dict and fn_dict['α'] is not None:
            data['α'] = fn_dict['α']
            cutval = "alphacut"
        elif 'α-α' in fn_dict:  # Fallback to 'α-α' if 'α' is missing/None
            data['α-α'] = fn_dict['α-α']
            cutval = "alphadash"

        # Add the fn columns
        data.update({
            'fn₁ (Y̲,ρ̲)': fn_dict['fn1'],
            'fn₂ (Y̅,ρ̲)': fn_dict['fn2'],
            'fn₃ (Y̲,ρ̅ )': fn_dict['fn3'], 
            'fn₄ (Y̅,ρ̅ )': fn_dict['fn4']
        })

        # Create DataFrame if either α or α-α exists
        # if fn_dict['α'] or fn_dict['α-α']:
        df = pd.DataFrame(data)

        title_row = 1        
        title_cell = ws.cell(row=title_row, column=1, value=filetitle)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
        # Title row
        title_row = 2
        if(cutval == "alphacut"):
            title_cell = ws.cell(row=title_row, column=1, value="Table : α-Cut")
        else:
            title_cell = ws.cell(row=title_row, column=1, value="Table : α-α'-Cut")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)

         # Title row
        title_row = 8
        if(cutval == "alphacut"):
            title_cell = ws.cell(row=title_row, column=1, value="Bar Chart : α-Cut")
        else:
            title_cell = ws.cell(row=title_row, column=1, value="Bar Chart : α-α'-Cut")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)
        # Insert DataFrame starting from a free area (e.g., row 32)
        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                # ws.cell(row=r_idx, column=c_idx, value=value)
                # ws.cell.alignment = Alignment(horizontal='right')
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='right')  # ✅ This is correct

        # Save workbook
        wb.save(output_excel)
        print(f"Successfully created: {os.path.abspath(output_excel)}")

    except Exception as e:
        print(f"Error: {str(e)}")
    finally:
        # Clean up temp files
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as e:
                print(f"Warning: Could not delete temp file {temp_file}: {str(e)}")


def export_images_to_excel2(img1, img2, img3, img4, output_excel="output.xlsx", max_width=500, max_height=300,fn_dict='',fn_dict_dash='',filetitle=''):
    """
    Export two resized PNG images to an Excel file with proper temp file handling.
    """
    temp_files = []  # To keep track of temporary files
    
    fn_dict_str = fn_dict
    fn_dict = json.loads(fn_dict_str.replace("'", '"'))  # Replace single quotes with double quotes 
    fn_dict_str2 = fn_dict_dash
    fn_dict_dash = json.loads(fn_dict_str2.replace("'", '"')) 

    # print(type(fn_dict))
    # print(fn_dict)
    try:
        # Validate input images
        for img_path in [img1, img2, img3, img4]:
            if not os.path.exists(img_path):
                raise FileNotFoundError(f"Image file not found: {img_path}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Line Chart"

        # Function to resize image and return temp path
        def resize_image(img_path):
            img = PILImage.open(img_path)
            img.thumbnail((max_width, max_height))            
            # Create temp file in system's temp directory
            fd, temp_path = tempfile.mkstemp(suffix='.png')
            os.close(fd)  # We only need the path
            img.save(temp_path)
            temp_files.append(temp_path)  # Track for cleanup
            return temp_path

        # Process images
        resized_img1 = resize_image(img1)
        resized_img2 = resize_image(img2)

        resized_img3 = resize_image(img3)
        resized_img4 = resize_image(img4)

        # Add images to worksheet
        ws.add_image(Image(resized_img1), "A9")
        ws.add_image(Image(resized_img2), "A27")

        ws.add_image(Image(resized_img3), "A45")
        ws.add_image(Image(resized_img4), "A63")

        data = {}        
        cutval = ""
        if 'α' in fn_dict and fn_dict['α'] is not None:
            data['α'] = fn_dict['α']
            cutval = "alphacut"
        elif 'α-α' in fn_dict:  # Fallback to 'α-α' if 'α' is missing/None
            data['α-α'] = fn_dict['α-α']
            cutval = "alphadash"
        # Add the fn columns
        data.update({
            '(Y̲)': fn_dict['Ymin'],
            '(Y̅)': fn_dict['Ymax'],
            '(ρ̲)': fn_dict['dmin'], 
            '(ρ̅ )': fn_dict['dmax']
        })
        # Create DataFrame if either α or α-α exists
        # if fn_dict['α'] or fn_dict['α-α']:
        df = pd.DataFrame(data)

        # Second alpha-dash ::
        data2 = {}        
        cutval2 = ""
        if 'α' in fn_dict_dash and fn_dict_dash['α'] is not None:
            data2['α'] = fn_dict_dash['α']
            cutval2 = "alphacut"
        elif 'α-α' in fn_dict_dash:  # Fallback to 'α-α' if 'α' is missing/None
            data2['α-α'] = fn_dict_dash['α-α']
            cutval2 = "alphadash"
        # Add the fn columns
        data2.update({
            '(Y̲)': fn_dict_dash['Ymin'],
            '(Y̅)': fn_dict_dash['Ymax'],
            '(ρ̲)': fn_dict_dash['dmin'], 
            '(ρ̅ )': fn_dict_dash['dmax']
        })
        # Create DataFrame if either α or α-α exists
        # if fn_dict['α'] or fn_dict['α-α']:
        df2 = pd.DataFrame(data2)

        title_row = 1        
        title_cell = ws.cell(row=title_row, column=1, value=filetitle)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)

        # Title row
        title_row = 2
        if(cutval == "alphacut"):
            title_cell = ws.cell(row=title_row, column=1, value="Table : α-Cut")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)

        title_row = 2
        if(cutval2 == "alphadash"):
            title_cell = ws.cell(row=title_row, column=7, value="Table : α-α'-Cut")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=7, end_row=title_row, end_column=11)

         # Title row
        title_row = 8
        if(cutval == "alphacut"):
            title_cell = ws.cell(row=title_row, column=1, value="Line Chart : α-Cut")
        else:
            title_cell = ws.cell(row=title_row, column=1, value="Line Chart : α-α'-Cut")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)
        # Insert DataFrame starting from a free area (e.g., row 32)
        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                # ws.cell(row=r_idx, column=c_idx, value=value)
                # ws.cell.alignment = Alignment(horizontal='right')
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='right')  # ✅ This is correct
        
        for r_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=7):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='right')  # ✅ This is correct

        # Save workbook
        wb.save(output_excel)
        print(f"Successfully created: {os.path.abspath(output_excel)}")

    except Exception as e:
        print(f"Error: {str(e)}")
    finally:
        # Clean up temp files
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as e:
                print(f"Warning: Could not delete temp file {temp_file}: {str(e)}")

# def export_images_to_excel(img1, img2, output_excel="output.xlsx", max_width=500, max_height=300):
#     """
#     Export two resized PNG images to an Excel file with proper temp file handling.
#     """
#     temp_files = []  # To keep track of temporary files
    
#     try:
#         # Validate input images
#         for img_path in [img1, img2]:
#             if not os.path.exists(img_path):
#                 raise FileNotFoundError(f"Image file not found: {img_path}")

#         # Create workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Images"

#         # Function to resize image and return temp path
#         def resize_image(img_path):
#             img = PILImage.open(img_path)
#             img.thumbnail((max_width, max_height))            
#             # Create temp file in system's temp directory
#             fd, temp_path = tempfile.mkstemp(suffix='.png')
#             os.close(fd)  # We only need the path
#             img.save(temp_path)
#             temp_files.append(temp_path)  # Track for cleanup
#             return temp_path

#         # Process images
#         resized_img1 = resize_image(img1)
#         resized_img2 = resize_image(img2)

#         # Add images to worksheet
#         ws.add_image(Image(resized_img1), "A1")
#         ws.add_image(Image(resized_img2), "A16")

#         # Adjust row heights to accommodate images
#         # ws.row_dimensions[1].height = max_height * 0.75
#         # ws.row_dimensions[20].height = max_height * 0.75

#         # Save workbook
#         wb.save(output_excel)
#         print(f"Successfully created: {os.path.abspath(output_excel)}")

#     except Exception as e:
#         print(f"Error: {str(e)}")
#     finally:
#         # Clean up temp files
#         for temp_file in temp_files:
#             try:
#                 if os.path.exists(temp_file):
#                     os.remove(temp_file)
#             except Exception as e:
#                 print(f"Warning: Could not delete temp file {temp_file}: {str(e)}")




# Example Usage
# export_images_to_excel(
#     img1="static/img/Alpha_Table.png",
#     img2="static/img/Bar_alpha_dashMF.png",
#     output_excel="images_exported.xlsx",
#     max_width=600,  # Custom max width
#     max_height=400  # Custom max height
# )